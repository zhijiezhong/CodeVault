# -*- coding: utf-8 -*-
# @Time    : 2023/11/20
# @Author  : Zhong Zhijie

import logging
import os
import random
import sys
from collections import defaultdict

import numpy as np
import torch
from matplotlib import pyplot as plt
from texttable import Texttable
from torch.nn import Module

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def set_seed(seed: int) -> None:
    """
    设置随机种子，使得结果可以复现
    :param seed: 随机种子
    :return: None
    """
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True


def seed_everything(seed_value: int) -> None:
    """
    设置随机种子
    :param seed_value: 随机种子
    :return: None
    """
    random.seed(seed_value)
    np.random.seed(seed_value)
    torch.manual_seed(seed_value)
    os.environ['PYTHONHASHSEED'] = str(seed_value)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed_value)
        torch.cuda.manual_seed_all(seed_value)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = True


def print_config_summary(args, config=None):
    def print_config(args):
        """
        Prints out options and arguments.

        Parameters
        ----------
        args : dict or argparse.Namespace
            Configuration options either as a dictionary or argparse Namespace.
        """
        # Yaml config is a dictionary while parser arguments is an object. Use vars() only on parser arguments.
        if type(args) is not dict:
            args = vars(args)
        keys = args.keys()
        # Initialize table
        table = Texttable()
        # Add rows to the table under two columns ("Parameter", "Value").
        table.add_rows([["Parameter", "Value"]] + [[k.replace("_", " ").capitalize(), args[k]] for k in keys])
        # Print the table.
        print(table.draw())
        return table.draw()

    """Prints out summary of options and arguments used"""
    # Summarize config on the screen as a sanity check
    print(100 * "=")
    print(f"Arguments being used:\n")
    args_draw = print_config(args)
    print(100 * "=")
    if config is not None:
        print(f"Here is the configuration being used:\n")
        print_config(config)
        print(100 * "=")

    return args_draw


def mkdir(directory: str) -> None:
    """
    创建文件夹
    :param directory: 路径
    :return: None
    """
    if not os.path.exists(directory):
        os.makedirs(directory)


def save_model(model: Module, filepath: str) -> None:
    """
    保存 PyTorch模型的状态字典到指定的文件路径
    :param model: PyTorch模型实例
    :param filepath: 要保存模型的文件路径
    :return: None
    """
    torch.save(model.state_dict(), filepath)


def load_model(model: Module, filepath: str) -> Module:
    """
    加载保存的 PyTorch 模型状态字典到指定的模型实例中
    :param model: PyTorch 模型实例（空的，用于加载参数）
    :param filepath: 已保存模型的文件路径
    :return: 已加载模型参数的 PyTorch 模型实例
    """
    # 加载模型的状态字典到指定的模型实例中
    model.load_state_dict(torch.load(filepath))
    return model


# 已经有了日志类了，这个可以不使用了
def get_logger(name: str, save_dir: str, file_name: str):
    """
    日志
    :param name: logger名称
    :param save_dir: 日志保存的目录
    :param file_name: 日志保存的文件名
    :return: logger
    """
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)

    ch = logging.StreamHandler(stream=sys.stdout)
    ch.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s %(name)s %(levelname)s: %(message)s", datefmt='%Y/%m/%d %I:%M:%S')
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    if save_dir:
        fh = logging.FileHandler(os.path.join(save_dir, file_name))
        fh.setLevel(logging.INFO)
        fh.setFormatter(formatter)
        logger.addHandler(fh)

    return logger


def create_exp_folder(base_dir: str):
    """
    检查并创建按照序号递增的实验文件夹
    :param base_dir: 基础目录路径，用于存放实验文件夹
    :return: 创建的实验文件夹路径
    """
    exp_num = 1
    exp_folder = os.path.join(base_dir, f"exp{exp_num}")

    # 检查是否存在该实验文件夹，如果存在则递增序号直到找到未被占用的序号
    while os.path.exists(exp_folder):
        exp_num += 1
        exp_folder = os.path.join(base_dir, f"exp{exp_num}")

    # 创建实验文件夹
    os.makedirs(exp_folder)
    return exp_folder


def get_exp_sub_folder_path(output_dir):
    output_dir = create_exp_folder(output_dir)
    log_dir = os.path.join(output_dir, 'log')
    model_dir = os.path.join(output_dir, 'model')
    figures_dir = os.path.join(output_dir, 'figures')
    mkdir(log_dir)
    mkdir(model_dir)
    mkdir(figures_dir)
    return log_dir, model_dir, figures_dir


def plot_dict_to_individual_files(metric: defaultdict, folder_path: str):
    def set_sci_style():
        # 设置 Matplotlib 样式为 SCI 形式
        plt.style.use('seaborn-white')  # 使用白色背景
        plt.rcParams.update({
            'font.family': 'serif',
            'font.serif': ['Times New Roman']
        })

    # 创建文件夹保存图表（如果不存在）
    os.makedirs(os.path.join(folder_path, 'pdf'), exist_ok=True)
    os.makedirs(os.path.join(folder_path, 'png'), exist_ok=True)

    # # 应用科学风格样式
    # set_sci_style()

    # 遍历字典中的键值对，并绘制每个键值对对应的图表
    for key, value in metric.items():
        plt.figure(figsize=(8, 6))  # 设置图像尺寸

        plt.plot(value)
        plt.xlabel('Epochs')  # x轴标签
        plt.ylabel(key)  # y轴标签（使用键作为y轴标签）

        # 构建保存路径并保存为 PDF 文件
        pdf_save_path = os.path.join(folder_path, 'pdf', f'{key.lower().replace(" ", "_")}_plot.pdf')
        plt.savefig(pdf_save_path, format='pdf')  # 保存为 PDF 格式

        # 构建保存路径并保存为 PNG 文件
        png_save_path = os.path.join(folder_path, 'png', f'{key.lower().replace(" ", "_")}_plot.png')
        plt.savefig(png_save_path, format='png')  # 保存为 PNG 格式

        plt.close()  # 关闭当前图表


def extract_last_two_values(directory):
    max_test_auc_values = []
    max_test_ap_values = []
    max_values = {}

    # 遍历目录下的exp文件夹
    for exp_folder in os.listdir(directory):
        exp_path = os.path.join(directory, exp_folder)
        if os.path.isdir(exp_path):
            log_path = os.path.join(exp_path, 'log', '4.best_metric.log')

            # 检查log文件是否存在
            if os.path.exists(log_path):
                with open(log_path, 'r') as file:
                    lines = file.readlines()

                    # todo: 不同的评价指标所在位置不一样就需要更改此处
                    # 提取倒数第二行和倒数第一行的数值
                    test_auc = float(lines[-1].split()[-1][: -1])
                    test_ap = float(lines[-2].split()[-1][: -1])

                    max_test_auc_values.append(test_auc)
                    max_test_ap_values.append(test_ap)

                    max_values[test_auc] = (exp_folder, 'Test AUC')
                    max_values[test_ap] = (exp_folder, 'Test AP')

    max_test_auc = max(max_test_auc_values)
    max_test_ap = max(max_test_ap_values)

    # 获取最大值对应的exp文件夹和指标
    max_test_auc_exp, metric_auc = max_values[max_test_auc]
    max_test_ap_exp, metric_ap = max_values[max_test_ap]

    print("Max Test AUC:", (max_test_auc, max_test_auc_exp, metric_auc))
    print("Max Test AP:", (max_test_ap, max_test_ap_exp, metric_ap))


def write_to_excel(parameters: list, values: list, metric: dict, compare: list = None, verbose: bool = True,
                   filename: str = 'output.xlsx') -> None:
    """
    将参数和值及其对应的指标写入Excel文件，并设置字体为新罗马，居中对齐。

    :param parameters: 参数列表
    :param values: 参数值列表
    :param metric: 指标字典
    :param compare: 比较指标（默认为None）
    :parm verbose: 是否一起写入参数和参数值（默认为True）
    :param filename: 输出的Excel文件名（默认为'output.xlsx'）
    """

    def find_first_empty_column(ws):
        """找到工作表中第一行第一个两列都是空的列"""
        col = 1
        while True:
            if ws.cell(row=1, column=col).value is None and ws.cell(row=1, column=col + 1).value is None:
                return col
            col += 1

    def auto_adjust_column_width(ws, start_col, end_col):
        """自动调整列宽"""
        for col in range(start_col, end_col + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    if compare is None:
        compare = [1] * len(metric)

    # 尝试加载现有工作簿，如果不存在则创建新的工作簿
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    # 找到第一个空的列
    start_column = find_first_empty_column(ws)

    # 如果不是第一列，则多添加一个空白列
    if start_column != 1:
        start_column += 1

    # 设置字体和居中对齐样式
    font = Font(name='Times New Roman')
    alignment = Alignment(horizontal='center', vertical='center')

    # 将参数和值作为单个单元格的值写入
    if verbose:
        parameters_pair = '_'.join(f'{param}: {value}' for param, value in zip(parameters, values))
    else:
        parameters_pair = '_'.join(values)
    cell = ws.cell(row=1, column=start_column, value=parameters_pair)
    cell.font = font
    cell.alignment = alignment

    # 写入指标
    start_column += 1
    for idx, (key, values) in enumerate(metric.items()):
        # 根据compare数组判断是最大值还是最小值
        max_index = values.index(max(values)) if compare[idx] == 1 else values.index(min(values))

        tmp = {k: v[max_index] for k, v in metric.items()}

        cell = ws.cell(row=1, column=start_column + idx, value='Max_' + key)
        cell.font = font
        cell.alignment = alignment
        cell = ws.cell(row=2, column=start_column + idx, value=str(tmp))
        cell.font = font
        cell.alignment = alignment

    # 自动调整列宽
    auto_adjust_column_width(ws, 1, start_column + len(metric) - 1)

    wb.save(filename)
    print(f"Data written to {filename} successfully.")
